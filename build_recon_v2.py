"""
Nalanda Enterprises vs Imagine Marketing India (boAt)
Credit Note Reconciliation FY 2025-26 — Version 2
Includes: Primary Billing (Tally), NLC Scheme Rates, Fuzzy Matching, CN Statement
"""

import openpyxl
from pyxlsb import open_workbook as open_xlsb
from datetime import datetime
from collections import defaultdict
import re
import difflib
import csv, io
import xlsxwriter

SCRATCHPAD = '/tmp/claude-0/-home-user-myaiproject/f4a54008-4277-5876-b3ea-4bf7e250e360/scratchpad/'
UPLOADS = '/root/.claude/uploads/f4a54008-4277-5876-b3ea-4bf7e250e360/'
OUTPUT_FILE = SCRATCHPAD + 'Nalanda_IML_SchemeRecon_v2.xlsx'

MONTHS_ORDER = ['Apr-2025','May-2025','Jun-2025','Jul-2025','Aug-2025','Sep-2025',
                'Oct-2025','Nov-2025','Dec-2025','Jan-2026','Feb-2026','Mar-2026']

NLC_FILES = {
    'Apr-2025': (UPLOADS + '9a3d20d4-Price_with_SchemeApr25.xlsb', 'xlsb', 'audio'),
    'May-2025': (UPLOADS + '275d65d4-GT_price_list_with_SchemeMay25Prateek_.xlsx', 'xlsx', 'gt'),
    'Jun-2025': (UPLOADS + 'bf0c515c-Price_with_SchemeJun25.xlsb', 'xlsb', 'audio'),
    'Jul-2025': (UPLOADS + '23a8e68b-Price_with_SchemeJul25.xlsb', 'xlsb', 'audio'),
    'Aug-2025': (UPLOADS + '8b9a0cc7-Price_with_SchemeAug25.xlsb', 'xlsb', 'audio'),
    'Sep-2025': (UPLOADS + '9a60d04d-Price_with_SchemeSep25.xlsb', 'xlsb', 'audio'),
    'Oct-2025': (UPLOADS + '43972c0b-Price_with_SchemeOct25.xlsb', 'xlsb', 'audio'),
    'Nov-2025': (UPLOADS + '5d8182b9-Price_with_SchemeNov25.xlsb', 'xlsb', 'audio'),
    'Dec-2025': (UPLOADS + '1536f02a-Price_with_SchemeDec25.xlsb', 'xlsb', 'audio'),
    'Jan-2026': (UPLOADS + '2948d766-Price_with_SchemeJan26.xlsb', 'xlsb', 'audio'),
    'Feb-2026': (UPLOADS + '730b5869-Feb_Price_with_Scheme.xlsx', 'xlsx', 'feb'),
    'Mar-2026': (UPLOADS + 'c11aec26-GTDB_NLCSharedMar26.xlsx', 'xlsx', 'mar'),
}

STATEMENT_FILE = UPLOADS + '623f85c7-Statement_0001100164.xlsx_13.xls'
PRIMARY_FILE   = '/tmp/imagine27.xlsx'

# ─── COLOR / NOISE WORDS TO STRIP DURING FUZZY MATCH ──────────────────────
COLOR_WORDS = {
    'black','white','blue','red','green','grey','gray','gold','silver','rose',
    'midnight','navy','teal','yellow','orange','purple','pink','brown','beige',
    'turquoise','azure','cobalt','indigo','violet','maroon','cream','ivory',
    'champagne','bronze','charcoal','slate','carbon','matte','glossy','gloss',
    'metal','metallic','active','pro','plus','lite','mini','max','ultra','neo',
    'prime','classic','elite','premium','smart','v2','gen2','anc','nc',
    'blazing','comet','ash','coral','galaxy','cosmic','onyx','ebony',
    'cherry','blossom','dusk','zinc','sand','pearl','chrome','titanium',
    'raspberry','lime','mint','aqua','mahogany','olive','cyan','magenta',
    'platinum','pewter','gunmetal','heather','lavender','lilac','fuchsia',
    'haze','smoky','storm','lightning','thunder','arctic','aurora','forest',
    'jungle','ocean','sky','sunrise','sunset','twilight','dusk','dawn',
    'black+', 'white+', 'neon', 'edition', 'se',
}

BRAND_PREFIXES = ['boat ', 'baot ', 'boa ', 'bo at ', 'imagine ', 'powerbank ']

def clean_name(name):
    """Normalize product name for fuzzy matching."""
    if not name:
        return ''
    n = str(name).lower().strip()
    # Remove brand prefixes
    for bp in BRAND_PREFIXES:
        if n.startswith(bp):
            n = n[len(bp):]
    # Remove leading/trailing noise
    n = re.sub(r'\s+', ' ', n)
    # Remove color words that are standalone tokens
    tokens = n.split()
    tokens = [t for t in tokens if t not in COLOR_WORDS and not re.match(r'^[a-z]$', t)]
    return ' '.join(tokens).strip()

def word_overlap_score(a, b):
    """Score based on word overlap between two cleaned names."""
    wa = set(a.split())
    wb = set(b.split())
    if not wa or not wb:
        return 0
    intersection = wa & wb
    return len(intersection) / max(len(wa), len(wb))

def best_match(tally_name, nlc_names_clean, nlc_names_orig, threshold=0.4):
    """Find best NLC match for a tally product name."""
    tc = clean_name(tally_name)
    if not tc:
        return None, 0
    best_score = 0
    best_orig = None
    for nc, no in zip(nlc_names_clean, nlc_names_orig):
        score = word_overlap_score(tc, nc)
        if score > best_score:
            best_score = score
            best_orig = no
    # Also try difflib sequence matcher as fallback
    if best_score < threshold:
        matches = difflib.get_close_matches(tc, nlc_names_clean, n=1, cutoff=0.5)
        if matches:
            idx = nlc_names_clean.index(matches[0])
            seq_score = difflib.SequenceMatcher(None, tc, nlc_names_clean[idx]).ratio()
            if seq_score > best_score:
                best_score = seq_score
                best_orig = nlc_names_orig[idx]
    return best_orig, best_score

# ─── READ NLC FILES ───────────────────────────────────────────────────────
def read_nlc_xlsb_audio(path):
    """Read xlsb NLC file (audio format): FamilyType | Name | DistPrice | Sec | Pri | NLC"""
    products = {}
    with open_xlsb(path) as wb:
        with wb.get_sheet(1) as ws:
            header_found = False
            for row in ws.rows():
                vals = [c.v for c in row]
                if not any(v for v in vals):
                    continue
                if vals[0] == 'Product Family Type' or vals[0] == 'Product Family Type':
                    header_found = True
                    continue
                if header_found and vals[1] and isinstance(vals[1], str):
                    name = str(vals[1]).strip()
                    if name.lower() in ('product family name-final -a', 'note', ''):
                        continue
                    try:
                        sec = float(vals[3]) if vals[3] else 0
                        pri = float(vals[4]) if vals[4] else 0
                        dist_price = float(vals[2]) if vals[2] else 0
                        nlc_val = float(vals[5]) if vals[5] else 0
                        ftype = str(vals[0]).strip() if vals[0] else ''
                    except (TypeError, ValueError):
                        continue
                    products[name] = {
                        'family_type': ftype,
                        'dist_price': dist_price,
                        'secondary': sec,
                        'primary': pri,
                        'nlc': nlc_val,
                    }
    return products

def read_nlc_xlsx_gt(path):
    """Read GT xlsx NLC file (May-2025): MCode | MName | Brand | PFamily | PFamilyType | ... | Name | Price | Sec | Pri | MaxAddl | NLC"""
    products = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == 'M Code':
            header_row = i
            break

    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2.active
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        if not row[0]:
            continue
        try:
            name = str(row[6]).strip() if row[6] and str(row[6]).strip() != '-' else str(row[3]).strip()
            sec = float(row[8]) if row[8] else 0
            pri = float(row[9]) if row[9] else 0
            dist_price = float(row[7]) if row[7] else 0
            nlc_val = float(row[11]) if row[11] else 0
            ftype = str(row[4]).strip() if row[4] else ''
        except (TypeError, ValueError):
            continue
        if name and name != '-':
            products[name] = {
                'family_type': ftype,
                'dist_price': dist_price,
                'secondary': sec,
                'primary': pri,
                'nlc': nlc_val,
            }
    return products

def read_nlc_xlsx_feb(path):
    """Feb-2026: ProductFamily | MatCode | Brand | SubCat | MatDesc | BillingGST | W/OGST | Sec | Pri | Addl | TOT | FinalNLC"""
    products = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_idx = None
    rows_list = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows_list):
        if row[0] == 'Product Family' or (row[2] and str(row[2]).strip() == 'Brand'):
            header_idx = i
            break
    if header_idx is None:
        return products
    for row in rows_list[header_idx+1:]:
        if not row[0]:
            continue
        try:
            name = str(row[0]).strip()  # Product Family as name
            # Also check material description
            mat_desc = str(row[4]).strip() if row[4] else ''
            sec = float(row[7]) if row[7] else 0
            pri = float(row[8]) if row[8] else 0
            addl = float(row[9]) if row[9] else 0
            dist_price = float(row[5]) if row[5] else 0
            nlc_val = float(row[11]) if row[11] else 0
            ftype = str(row[3]).strip() if row[3] else ''
        except (TypeError, ValueError):
            continue
        if name:
            products[name] = {
                'family_type': ftype,
                'dist_price': dist_price,
                'secondary': sec,
                'primary': pri,
                'additional': addl,
                'nlc': nlc_val,
                'mat_desc': mat_desc,
            }
    return products

def read_nlc_xlsx_mar(path):
    """Mar-2026: blank | ProductFamily | MatCode | Brand | SubCat | MatDesc | BillingGST | Basic | Sec | Pri | Addl | FinalNLC | Remarks"""
    products = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_idx = None
    rows_list = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows_list):
        if row[1] == 'Product Family' or (len(row) > 2 and row[2] == 'Material Code'):
            header_idx = i
            break
    if header_idx is None:
        return products
    for row in rows_list[header_idx+1:]:
        if not row[1]:
            continue
        try:
            name = str(row[1]).strip()
            mat_desc = str(row[5]).strip() if row[5] else ''
            sec = float(row[8]) if row[8] else 0
            pri = float(row[9]) if row[9] else 0
            addl = float(row[10]) if row[10] else 0
            dist_price = float(row[6]) if row[6] else 0
            nlc_val = float(row[11]) if row[11] else 0
            ftype = str(row[3]).strip() if row[3] else ''
        except (TypeError, ValueError):
            continue
        if name:
            products[name] = {
                'family_type': ftype,
                'dist_price': dist_price,
                'secondary': sec,
                'primary': pri,
                'additional': addl,
                'nlc': nlc_val,
                'mat_desc': mat_desc,
            }
    return products

def load_all_nlc():
    """Load all NLC files, return dict: month -> {product_name -> scheme_dict}"""
    nlc_data = {}
    for month, (path, fmt, style) in NLC_FILES.items():
        try:
            if fmt == 'xlsb':
                products = read_nlc_xlsb_audio(path)
            elif style == 'gt':
                products = read_nlc_xlsx_gt(path)
            elif style == 'feb':
                products = read_nlc_xlsx_feb(path)
            elif style == 'mar':
                products = read_nlc_xlsx_mar(path)
            else:
                products = {}
            nlc_data[month] = products
            print(f"  NLC {month}: {len(products)} products loaded")
        except Exception as e:
            print(f"  NLC {month}: ERROR {e}")
            nlc_data[month] = {}
    return nlc_data

# ─── READ PRIMARY BILLING (TALLY IMAGINE_27) ─────────────────────────────
def load_primary_billing():
    """Returns: month -> list of {invoice, date, product, qty, rate, value}"""
    billing = defaultdict(list)
    wb = openpyxl.load_workbook(PRIMARY_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    cur_date = None
    cur_invoice = None
    cur_total_qty = None
    cur_total_value = None

    # Also collect invoice-level totals
    invoices = []

    for row in rows:
        # Invoice header: col 0 = date, col 3 = invoice no, col 4 = qty, col 6 = value
        if isinstance(row[0], datetime):
            cur_date = row[0]
            cur_invoice = str(row[3]).strip() if row[3] else ''
            cur_total_qty = row[4]
            cur_total_value = row[6]
            invoices.append({
                'date': cur_date,
                'invoice': cur_invoice,
                'total_qty': cur_total_qty,
                'total_value': cur_total_value,
            })
        elif cur_date and row[1] and row[4]:
            # Product line
            name = str(row[1]).strip()
            # Skip header/metadata rows
            if name in ('Particulars', 'IMAGINE MARKETING  LTD ', 'NALANDA ENTERPRISES',
                        '1-Apr-25 to 31-Mar-26', 'Date', 'Grand Total'):
                continue
            qty = row[4]
            rate = row[5]
            value = row[6]
            if not isinstance(qty, (int, float)):
                continue
            month_key = cur_date.strftime('%b-%Y')
            billing[month_key].append({
                'date': cur_date,
                'invoice': cur_invoice,
                'product': name,
                'qty': float(qty),
                'rate': float(rate) if rate else 0,
                'value': float(value) if value else 0,
            })

    return billing, invoices

# ─── READ STATEMENT (CN DATA) ────────────────────────────────────────────
def load_statement():
    """Returns list of dicts for all transactions from IML statement."""
    records = []
    with open(STATEMENT_FILE, 'rb') as f:
        raw = f.read()
    text = raw.decode('utf-16')
    reader = csv.reader(io.StringIO(text), delimiter='\t')
    rows = list(reader)
    header = None
    for row in rows:
        if not row:
            continue
        if row[0].strip() == 'Customer Name':
            header = [c.strip() for c in row]
            continue
        if header is None:
            continue
        if len(row) < len(header):
            row += [''] * (len(header) - len(row))
        rec = {h: v.strip() for h, v in zip(header, row)}
        records.append(rec)
    return records, header

def parse_amount(s):
    if not s:
        return 0.0
    s = s.replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_date_stmt(s):
    s = s.strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

# ─── MAIN RECONCILIATION ─────────────────────────────────────────────────
print("=" * 70)
print("NALANDA ENTERPRISES — IML CREDIT NOTE RECONCILIATION FY 2025-26 v2")
print("=" * 70)

print("\n[1] Loading NLC Scheme Files...")
nlc_data = load_all_nlc()

print("\n[2] Loading Primary Billing (Tally)...")
billing, invoices = load_primary_billing()
for m in MONTHS_ORDER:
    lines = billing.get(m, [])
    total_qty = sum(l['qty'] for l in lines)
    total_val = sum(l['value'] for l in lines)
    print(f"  {m}: {len(lines)} product lines, {total_qty:.0f} qty, ₹{total_val:,.0f}")

print("\n[3] Loading Statement Data...")
stmt_records, stmt_header = load_statement()
CN_TYPES = {'DG', 'CN', 'ZC'}
RV_TYPES = {'RV'}

cn_records = [r for r in stmt_records if r.get('Document Type','').strip() in CN_TYPES]
rv_records = [r for r in stmt_records if r.get('Document Type','').strip() in RV_TYPES]
print(f"  Total records: {len(stmt_records)}")
print(f"  Credit Notes (DG/CN/ZC): {len(cn_records)}")
print(f"  Purchases (RV): {len(rv_records)}")

# Group CNs by month
cn_by_month = defaultdict(list)
for r in cn_records:
    d = parse_date_stmt(r.get('Document Date',''))
    if d:
        mk = d.strftime('%b-%Y')
        amt = parse_amount(r.get('Amount in LC',''))
        cn_by_month[mk].append({'date': d, 'doc_no': r.get('Document Number',''),
                                  'doc_type': r.get('Document Type',''),
                                  'amount': amt, 'assignment': r.get('Assignment',''),
                                  'ref': r.get('Reference Key 3','')})

# Group purchases by month
purch_by_month = defaultdict(list)
for r in rv_records:
    d = parse_date_stmt(r.get('Document Date',''))
    if d:
        mk = d.strftime('%b-%Y')
        amt = parse_amount(r.get('Amount in LC',''))
        purch_by_month[mk].append({'date': d, 'doc_no': r.get('Document Number',''), 'amount': amt})

print(f"\n[4] Building Product-Scheme Mapping (Fuzzy Match)...")

# For each month, match tally products to NLC products and compute expected scheme
# Results: month -> product -> {matched_nlc, qty, secondary_total, primary_total, match_score, match_conf}
month_product_scheme = {}
unmatched_products = defaultdict(set)

for month in MONTHS_ORDER:
    lines = billing.get(month, [])
    nlc = nlc_data.get(month, {})

    if not nlc:
        month_product_scheme[month] = []
        continue

    # Prepare NLC lookup
    nlc_names_orig = list(nlc.keys())
    nlc_names_clean = [clean_name(n) for n in nlc_names_orig]

    # Aggregate qty by product within this month
    prod_qty = defaultdict(float)
    prod_value = defaultdict(float)
    for line in lines:
        prod_qty[line['product']] += line['qty']
        prod_value[line['product']] += line['value']

    results = []
    for tally_prod, qty in prod_qty.items():
        matched_nlc_name, score = best_match(tally_prod, nlc_names_clean, nlc_names_orig)

        if matched_nlc_name and score >= 0.35:
            scheme = nlc[matched_nlc_name]
            sec_total = qty * scheme['secondary']
            pri_total = qty * scheme['primary']
            addl_total = qty * scheme.get('additional', 0)
            conf = 'HIGH' if score >= 0.65 else 'MEDIUM' if score >= 0.45 else 'LOW'
        else:
            scheme = {}
            sec_total = 0
            pri_total = 0
            addl_total = 0
            conf = 'UNMATCHED'
            unmatched_products[month].add(tally_prod)

        results.append({
            'tally_product': tally_prod,
            'qty': qty,
            'billing_value': prod_value[tally_prod],
            'matched_nlc': matched_nlc_name or 'NO MATCH',
            'match_score': round(score, 3),
            'match_conf': conf,
            'family_type': scheme.get('family_type', ''),
            'dist_price': scheme.get('dist_price', 0),
            'secondary_per_unit': scheme.get('secondary', 0),
            'primary_per_unit': scheme.get('primary', 0),
            'secondary_total': round(sec_total, 2),
            'primary_total': round(pri_total, 2),
            'additional_total': round(addl_total, 2),
            'expected_scheme_total': round(sec_total + pri_total + addl_total, 2),
        })

    month_product_scheme[month] = results
    total_exp = sum(r['expected_scheme_total'] for r in results)
    matched_count = sum(1 for r in results if r['match_conf'] != 'UNMATCHED')
    print(f"  {month}: {len(results)} products, {matched_count} matched, Expected Scheme ₹{total_exp:,.0f}")

print("\n[5] Building Summary Tables...")

# Monthly summary
monthly_summary = []
for month in MONTHS_ORDER:
    results = month_product_scheme.get(month, [])

    # Billing from Tally
    tally_qty = sum(r['qty'] for r in results)
    tally_value = sum(r['billing_value'] for r in results)

    # Expected scheme
    exp_secondary = sum(r['secondary_total'] for r in results)
    exp_primary = sum(r['primary_total'] for r in results)
    exp_total = sum(r['expected_scheme_total'] for r in results)

    # Actual CNs from statement
    cns = cn_by_month.get(month, [])
    # Negative amounts are credits
    act_cn = sum(abs(c['amount']) for c in cns if c['amount'] < 0)
    # Also count positive CN reversals
    act_cn_reversal = sum(c['amount'] for c in cns if c['amount'] > 0)
    net_cn = act_cn - act_cn_reversal

    # Purchases from statement
    purch = purch_by_month.get(month, [])
    stmt_purchase = sum(p['amount'] for p in purch if p['amount'] > 0)

    diff = net_cn - exp_total

    monthly_summary.append({
        'month': month,
        'tally_qty': tally_qty,
        'tally_billing_value': round(tally_value, 2),
        'stmt_purchase_value': round(stmt_purchase, 2),
        'exp_secondary_cn': round(exp_secondary, 2),
        'exp_primary_cn': round(exp_primary, 2),
        'exp_total_cn': round(exp_total, 2),
        'actual_cn_received': round(net_cn, 2),
        'cn_difference': round(diff, 2),
        'cn_count': len(cns),
        'status': 'EXCESS' if diff > 500 else ('SHORT' if diff < -500 else 'MATCHED'),
    })

print("\n=== MONTHLY SUMMARY ===")
print(f"{'Month':<12} {'Tally Qty':>10} {'Tally Value':>14} {'Exp Secondary':>14} {'Exp Primary':>12} {'Exp Total':>12} {'Actual CN':>12} {'Diff':>10} {'Status':<10}")
for s in monthly_summary:
    print(f"{s['month']:<12} {s['tally_qty']:>10,.0f} {s['tally_billing_value']:>14,.0f} {s['exp_secondary_cn']:>14,.0f} {s['exp_primary_cn']:>12,.0f} {s['exp_total_cn']:>12,.0f} {s['actual_cn_received']:>12,.0f} {s['cn_difference']:>10,.0f} {s['status']:<10}")

total_exp_all = sum(s['exp_total_cn'] for s in monthly_summary)
total_act_all = sum(s['actual_cn_received'] for s in monthly_summary)
print(f"\nFY TOTAL: Expected ₹{total_exp_all:,.0f} | Actual ₹{total_act_all:,.0f} | Diff ₹{total_act_all - total_exp_all:,.0f}")

# ─── BUILD EXCEL OUTPUT ────────────────────────────────────────────────────
print("\n[6] Building Excel Output...")

wb = xlsxwriter.Workbook(OUTPUT_FILE)

# ── FORMATS ──
fmt_title   = wb.add_format({'bold': True, 'font_size': 16, 'font_color': '#FFFFFF',
                               'bg_color': '#1F3864', 'align': 'center', 'valign': 'vcenter'})
fmt_hdr     = wb.add_format({'bold': True, 'bg_color': '#1F3864', 'font_color': '#FFFFFF',
                               'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
fmt_hdr2    = wb.add_format({'bold': True, 'bg_color': '#2E75B6', 'font_color': '#FFFFFF',
                               'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
fmt_hdr3    = wb.add_format({'bold': True, 'bg_color': '#70AD47', 'font_color': '#FFFFFF',
                               'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
fmt_num     = wb.add_format({'num_format': '#,##0.00', 'border': 1})
fmt_int     = wb.add_format({'num_format': '#,##0', 'border': 1})
fmt_pct     = wb.add_format({'num_format': '0.0%', 'border': 1})
fmt_text    = wb.add_format({'border': 1, 'text_wrap': True})
fmt_text_c  = wb.add_format({'border': 1, 'align': 'center'})
fmt_date    = wb.add_format({'num_format': 'dd-mmm-yy', 'border': 1})
fmt_month   = wb.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1, 'align': 'center'})
fmt_subtot  = wb.add_format({'bold': True, 'bg_color': '#FCE4D6', 'border': 1, 'num_format': '#,##0.00'})
fmt_subtot_int = wb.add_format({'bold': True, 'bg_color': '#FCE4D6', 'border': 1, 'num_format': '#,##0'})
fmt_grandtot = wb.add_format({'bold': True, 'bg_color': '#FFC000', 'border': 1, 'num_format': '#,##0.00', 'font_size': 11})
fmt_grandtot_t = wb.add_format({'bold': True, 'bg_color': '#FFC000', 'border': 1, 'font_size': 11})
fmt_green   = wb.add_format({'bg_color': '#C6EFCE', 'font_color': '#276221', 'border': 1, 'align': 'center', 'bold': True})
fmt_red     = wb.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'border': 1, 'align': 'center', 'bold': True})
fmt_orange  = wb.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C6500', 'border': 1, 'align': 'center', 'bold': True})
fmt_high    = wb.add_format({'bg_color': '#C6EFCE', 'border': 1, 'align': 'center'})
fmt_med     = wb.add_format({'bg_color': '#FFEB9C', 'border': 1, 'align': 'center'})
fmt_low     = wb.add_format({'bg_color': '#FFC7CE', 'border': 1, 'align': 'center'})
fmt_unm     = wb.add_format({'bg_color': '#D9D9D9', 'border': 1, 'align': 'center'})

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 1: MASTER DASHBOARD                                  ║
# ╚══════════════════════════════════════════════════════════════╝
ws1 = wb.add_worksheet('MASTER DASHBOARD')
ws1.set_tab_color('#1F3864')
ws1.set_column('A:A', 14)
ws1.set_column('B:B', 14)
ws1.set_column('C:C', 16)
ws1.set_column('D:D', 16)
ws1.set_column('E:E', 16)
ws1.set_column('F:F', 16)
ws1.set_column('G:G', 16)
ws1.set_column('H:H', 16)
ws1.set_column('I:I', 12)
ws1.set_column('J:J', 12)
ws1.set_row(0, 50)

ws1.merge_range('A1:J1', 'NALANDA ENTERPRISES vs IMAGINE MARKETING INDIA — SCHEME CREDIT NOTE RECONCILIATION FY 2025-26', fmt_title)
ws1.merge_range('A2:J2', 'Customer Code: 1100164 | Prepared: ' + datetime.now().strftime('%d-%b-%Y'),
                wb.add_format({'italic': True, 'align': 'center', 'bg_color': '#BDD7EE'}))

# KPI Row
ws1.set_row(3, 40)
kpis = [
    ('TOTAL PURCHASES\n(Tally)', f"₹{sum(s['tally_billing_value'] for s in monthly_summary)/1e7:.2f} Cr"),
    ('TOTAL CN\nEXPECTED', f"₹{sum(s['exp_total_cn'] for s in monthly_summary)/1e5:.2f} L"),
    ('TOTAL CN\nRECEIVED', f"₹{sum(s['actual_cn_received'] for s in monthly_summary)/1e5:.2f} L"),
    ('DIFFERENCE\n(Act - Exp)', f"₹{(sum(s['actual_cn_received'] for s in monthly_summary) - sum(s['exp_total_cn'] for s in monthly_summary))/1e5:.2f} L"),
    ('TOTAL INVOICES\n(Tally)', str(len(invoices))),
]
kpi_colors = ['#2E75B6', '#70AD47', '#FFC000', '#FF0000', '#7030A0']
for i, (label, value) in enumerate(kpis):
    fmt_kpi_lbl = wb.add_format({'bold': True, 'bg_color': kpi_colors[i], 'font_color': '#FFFFFF',
                                   'align': 'center', 'valign': 'top', 'text_wrap': True, 'font_size': 9})
    fmt_kpi_val = wb.add_format({'bold': True, 'bg_color': kpi_colors[i], 'font_color': '#FFFFFF',
                                   'align': 'center', 'valign': 'bottom', 'font_size': 14})
    col = i * 2
    ws1.merge_range(3, col, 3, col+1, label, fmt_kpi_lbl)
    ws1.merge_range(4, col, 4, col+1, value, fmt_kpi_val)

# Monthly Summary Table
ws1.set_row(6, 35)
headers_dash = ['Month', 'Tally\nQty', 'Tally Billing\nValue (₹)',
                'Stmt Purchase\n(₹)', 'Exp Secondary\nCN (₹)',
                'Exp Primary\nCN (₹)', 'Exp Total\nCN (₹)',
                'Actual CN\nReceived (₹)', 'Difference\n(₹)', 'Status']
for ci, h in enumerate(headers_dash):
    ws1.write(6, ci, h, fmt_hdr)

for ri, s in enumerate(monthly_summary):
    r = ri + 7
    ws1.write(r, 0, s['month'], fmt_month)
    ws1.write(r, 1, s['tally_qty'], fmt_int)
    ws1.write(r, 2, s['tally_billing_value'], fmt_num)
    ws1.write(r, 3, s['stmt_purchase_value'], fmt_num)
    ws1.write(r, 4, s['exp_secondary_cn'], fmt_num)
    ws1.write(r, 5, s['exp_primary_cn'], fmt_num)
    ws1.write(r, 6, s['exp_total_cn'], fmt_num)
    ws1.write(r, 7, s['actual_cn_received'], fmt_num)
    ws1.write(r, 8, s['cn_difference'], fmt_num)
    sfmt = fmt_green if s['status'] == 'MATCHED' else (fmt_red if s['status'] == 'SHORT' else fmt_orange)
    ws1.write(r, 9, s['status'], sfmt)

# Grand Total row
gr = len(monthly_summary) + 7
ws1.write(gr, 0, 'FY TOTAL', fmt_grandtot_t)
ws1.write(gr, 1, sum(s['tally_qty'] for s in monthly_summary), fmt_grandtot)
ws1.write(gr, 2, sum(s['tally_billing_value'] for s in monthly_summary), fmt_grandtot)
ws1.write(gr, 3, sum(s['stmt_purchase_value'] for s in monthly_summary), fmt_grandtot)
ws1.write(gr, 4, sum(s['exp_secondary_cn'] for s in monthly_summary), fmt_grandtot)
ws1.write(gr, 5, sum(s['exp_primary_cn'] for s in monthly_summary), fmt_grandtot)
ws1.write(gr, 6, sum(s['exp_total_cn'] for s in monthly_summary), fmt_grandtot)
ws1.write(gr, 7, sum(s['actual_cn_received'] for s in monthly_summary), fmt_grandtot)
total_diff = sum(s['actual_cn_received'] for s in monthly_summary) - sum(s['exp_total_cn'] for s in monthly_summary)
ws1.write(gr, 8, total_diff, fmt_grandtot)
ws1.write(gr, 9, 'NET DIFF', fmt_grandtot_t)

# Notes
nr = gr + 2
ws1.merge_range(nr, 0, nr, 9, '📌 IMPORTANT NOTES', fmt_hdr)
notes = [
    'Tally billing = Primary billing from Nalanda Tally ledger (IMAGINE MARKETING LTD MUMBAI 27) Apr\'25–Mar\'26',
    'Expected CN = Calculated from NLC scheme files (Secondary + Primary per unit × Tally quantity purchased)',
    'Actual CN = Credit notes received as per IML SAP statement (Doc Types: DG, CN, ZC)',
    'Product matching uses AI fuzzy logic (word overlap + sequence match). Confidence: HIGH ≥65%, MEDIUM ≥45%, LOW ≥35%',
    'UNMATCHED products = Tally products where no NLC scheme found; these are excluded from expected CN calculation',
    'Differences may arise due to: (1) Products not in scheme, (2) Scheme qty vs billing qty difference, (3) Delayed CNs',
    'See PRODUCT SCHEME DETAIL sheet for full product-wise scheme calculation',
    'See CN REGISTER sheet for all actual credit notes received',
]
for i, note in enumerate(notes):
    ws1.write(nr + 1 + i, 0, f'{i+1}. {note}', wb.add_format({'text_wrap': True, 'border': 1, 'valign': 'vcenter'}))
    ws1.merge_range(nr + 1 + i, 0, nr + 1 + i, 9, f'{i+1}. {note}',
                    wb.add_format({'text_wrap': True, 'border': 1, 'valign': 'vcenter'}))
    ws1.set_row(nr + 1 + i, 20)

print("  ✓ MASTER DASHBOARD")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 2: PRODUCT SCHEME DETAIL (Month-wise, Product-wise)  ║
# ╚══════════════════════════════════════════════════════════════╝
ws2 = wb.add_worksheet('PRODUCT SCHEME DETAIL')
ws2.set_tab_color('#70AD47')
ws2.set_column('A:A', 14)   # Month
ws2.set_column('B:B', 40)   # Tally Product
ws2.set_column('C:C', 35)   # Matched NLC Name
ws2.set_column('D:D', 15)   # Product Family Type
ws2.set_column('E:E', 10)   # Qty
ws2.set_column('F:F', 14)   # Billing Value
ws2.set_column('G:G', 12)   # Dist Price
ws2.set_column('H:H', 13)   # Sec Per Unit
ws2.set_column('I:I', 13)   # Pri Per Unit
ws2.set_column('J:J', 14)   # Sec Total
ws2.set_column('K:K', 14)   # Pri Total
ws2.set_column('L:L', 14)   # Exp Total Scheme
ws2.set_column('M:M', 12)   # Match Score
ws2.set_column('N:N', 12)   # Match Confidence
ws2.set_row(0, 35)

ws2.merge_range('A1:N1', 'PRODUCT-WISE SCHEME CALCULATION — ALL MONTHS', fmt_title)

hdrs2 = ['Month', 'Tally Product Name', 'Matched NLC Product Family', 'Family Type',
          'Qty\nPurchased', 'Billing Value\n(₹)', 'Dist Price\nwith Tax (₹)',
          'Secondary\nPer Unit (₹)', 'Primary\nPer Unit (₹)',
          'Secondary\nTotal (₹)', 'Primary\nTotal (₹)',
          'Total Expected\nScheme (₹)', 'Match\nScore', 'Match\nConfidence']
ws2.set_row(1, 35)
for ci, h in enumerate(hdrs2):
    ws2.write(1, ci, h, fmt_hdr)

row2 = 2
for month in MONTHS_ORDER:
    results = month_product_scheme.get(month, [])
    if not results:
        continue
    # Sort by expected scheme desc
    results_sorted = sorted(results, key=lambda x: -x['expected_scheme_total'])

    for r in results_sorted:
        cfmt = fmt_high if r['match_conf'] == 'HIGH' else \
               fmt_med  if r['match_conf'] == 'MEDIUM' else \
               fmt_low  if r['match_conf'] == 'LOW' else fmt_unm

        ws2.write(row2, 0, month, fmt_month)
        ws2.write(row2, 1, r['tally_product'], fmt_text)
        ws2.write(row2, 2, r['matched_nlc'], fmt_text)
        ws2.write(row2, 3, r['family_type'], fmt_text_c)
        ws2.write(row2, 4, r['qty'], fmt_int)
        ws2.write(row2, 5, r['billing_value'], fmt_num)
        ws2.write(row2, 6, r['dist_price'], fmt_num)
        ws2.write(row2, 7, r['secondary_per_unit'], fmt_num)
        ws2.write(row2, 8, r['primary_per_unit'], fmt_num)
        ws2.write(row2, 9, r['secondary_total'], fmt_num)
        ws2.write(row2, 10, r['primary_total'], fmt_num)
        ws2.write(row2, 11, r['expected_scheme_total'], fmt_num)
        ws2.write(row2, 12, r['match_score'], fmt_text_c)
        ws2.write(row2, 13, r['match_conf'], cfmt)
        row2 += 1

    # Subtotal for month
    sub_exp_sec = sum(r['secondary_total'] for r in results_sorted)
    sub_exp_pri = sum(r['primary_total'] for r in results_sorted)
    sub_exp_tot = sum(r['expected_scheme_total'] for r in results_sorted)
    sub_qty     = sum(r['qty'] for r in results_sorted)
    sub_val     = sum(r['billing_value'] for r in results_sorted)
    ws2.write(row2, 0, f'{month} TOTAL', fmt_grandtot_t)
    ws2.write(row2, 1, '', fmt_grandtot_t)
    ws2.write(row2, 2, '', fmt_grandtot_t)
    ws2.write(row2, 3, '', fmt_grandtot_t)
    ws2.write(row2, 4, sub_qty, fmt_grandtot)
    ws2.write(row2, 5, sub_val, fmt_grandtot)
    ws2.write(row2, 6, '', fmt_grandtot_t)
    ws2.write(row2, 7, '', fmt_grandtot_t)
    ws2.write(row2, 8, '', fmt_grandtot_t)
    ws2.write(row2, 9, sub_exp_sec, fmt_grandtot)
    ws2.write(row2, 10, sub_exp_pri, fmt_grandtot)
    ws2.write(row2, 11, sub_exp_tot, fmt_grandtot)
    ws2.write(row2, 12, '', fmt_grandtot_t)
    ws2.write(row2, 13, '', fmt_grandtot_t)
    row2 += 1

print("  ✓ PRODUCT SCHEME DETAIL")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 3: CN REGISTER (Actual CNs from Statement)          ║
# ╚══════════════════════════════════════════════════════════════╝
ws3 = wb.add_worksheet('CN REGISTER')
ws3.set_tab_color('#FF0000')
ws3.set_column('A:A', 12)
ws3.set_column('B:B', 14)
ws3.set_column('C:C', 10)
ws3.set_column('D:D', 20)
ws3.set_column('E:E', 14)
ws3.set_column('F:F', 30)
ws3.set_column('G:G', 14)
ws3.set_row(0, 35)

ws3.merge_range('A1:G1', 'CREDIT NOTE REGISTER — ACTUAL CNs FROM IML SAP STATEMENT (DG / CN / ZC)', fmt_title)

hdrs3 = ['CN Date', 'Month', 'Doc Type', 'CN Number', 'Amount (₹)', 'Assignment / Narration', 'Reference']
ws3.set_row(1, 25)
for ci, h in enumerate(hdrs3):
    ws3.write(1, ci, h, fmt_hdr)

row3 = 2
for month in MONTHS_ORDER:
    cns = cn_by_month.get(month, [])
    cns_sorted = sorted(cns, key=lambda x: x['date'])
    for cn in cns_sorted:
        ws3.write_datetime(row3, 0, cn['date'], fmt_date)
        ws3.write(row3, 1, month, fmt_month)
        ws3.write(row3, 2, cn['doc_type'], fmt_text_c)
        ws3.write(row3, 3, cn['doc_no'], fmt_text)
        ws3.write(row3, 4, cn['amount'], fmt_num)
        ws3.write(row3, 5, cn['assignment'], fmt_text)
        ws3.write(row3, 6, cn['ref'], fmt_text)
        row3 += 1

    if cns_sorted:
        sub_amt = sum(c['amount'] for c in cns_sorted)
        ws3.write(row3, 0, f'{month} TOTAL', fmt_grandtot_t)
        ws3.write(row3, 1, '', fmt_grandtot_t)
        ws3.write(row3, 2, f'{len(cns_sorted)} CNs', fmt_grandtot_t)
        ws3.write(row3, 3, '', fmt_grandtot_t)
        ws3.write(row3, 4, sub_amt, fmt_grandtot)
        ws3.write(row3, 5, '', fmt_grandtot_t)
        ws3.write(row3, 6, '', fmt_grandtot_t)
        row3 += 1

print("  ✓ CN REGISTER")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 4: TALLY PRIMARY BILLING                            ║
# ╚══════════════════════════════════════════════════════════════╝
ws4 = wb.add_worksheet('TALLY PRIMARY BILLING')
ws4.set_tab_color('#2E75B6')
ws4.set_column('A:A', 12)
ws4.set_column('B:B', 14)
ws4.set_column('C:C', 18)
ws4.set_column('D:D', 40)
ws4.set_column('E:E', 12)
ws4.set_column('F:F', 14)
ws4.set_column('G:G', 14)
ws4.set_row(0, 35)

ws4.merge_range('A1:G1', 'NALANDA ENTERPRISES — TALLY PRIMARY BILLING (IMAGINE MARKETING LTD MUMBAI 27) FY 2025-26', fmt_title)

hdrs4 = ['Date', 'Month', 'Invoice No.', 'Product Name', 'Qty', 'Rate (₹)', 'Value (₹)']
ws4.set_row(1, 25)
for ci, h in enumerate(hdrs4):
    ws4.write(1, ci, h, fmt_hdr)

row4 = 2
for month in MONTHS_ORDER:
    lines = billing.get(month, [])
    lines_sorted = sorted(lines, key=lambda x: (x['date'], x['invoice']))
    for line in lines_sorted:
        ws4.write_datetime(row4, 0, line['date'], fmt_date)
        ws4.write(row4, 1, month, fmt_month)
        ws4.write(row4, 2, line['invoice'], fmt_text)
        ws4.write(row4, 3, line['product'], fmt_text)
        ws4.write(row4, 4, line['qty'], fmt_int)
        ws4.write(row4, 5, line['rate'], fmt_num)
        ws4.write(row4, 6, line['value'], fmt_num)
        row4 += 1

    if lines_sorted:
        ws4.write(row4, 0, f'{month} TOTAL', fmt_grandtot_t)
        ws4.write(row4, 1, '', fmt_grandtot_t)
        ws4.write(row4, 2, f'{len(set(l["invoice"] for l in lines_sorted))} invoices', fmt_grandtot_t)
        ws4.write(row4, 3, '', fmt_grandtot_t)
        ws4.write(row4, 4, sum(l['qty'] for l in lines_sorted), fmt_grandtot)
        ws4.write(row4, 5, '', fmt_grandtot_t)
        ws4.write(row4, 6, sum(l['value'] for l in lines_sorted), fmt_grandtot)
        row4 += 1

print("  ✓ TALLY PRIMARY BILLING")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 5: NLC SCHEME REGISTER                              ║
# ╚══════════════════════════════════════════════════════════════╝
ws5 = wb.add_worksheet('NLC SCHEME REGISTER')
ws5.set_tab_color('#7030A0')
ws5.set_column('A:A', 12)
ws5.set_column('B:B', 35)
ws5.set_column('C:C', 20)
ws5.set_column('D:D', 14)
ws5.set_column('E:E', 12)
ws5.set_column('F:F', 12)
ws5.set_column('G:G', 12)
ws5.set_row(0, 35)

ws5.merge_range('A1:G1', 'NLC SCHEME REGISTER — ALL MONTHS (IML Scheme Letters)', fmt_title)
hdrs5 = ['Month', 'Product Family Name (NLC)', 'Product Family Type',
          'Dist. Price w/ Tax', 'Secondary\nPer Unit (₹)', 'Primary\nPer Unit (₹)', 'DB NLC (₹)']
ws5.set_row(1, 30)
for ci, h in enumerate(hdrs5):
    ws5.write(1, ci, h, fmt_hdr)

row5 = 2
for month in MONTHS_ORDER:
    nlc = nlc_data.get(month, {})
    for name, scheme in sorted(nlc.items()):
        ws5.write(row5, 0, month, fmt_month)
        ws5.write(row5, 1, name, fmt_text)
        ws5.write(row5, 2, scheme.get('family_type',''), fmt_text)
        ws5.write(row5, 3, scheme.get('dist_price', 0), fmt_num)
        ws5.write(row5, 4, scheme.get('secondary', 0), fmt_num)
        ws5.write(row5, 5, scheme.get('primary', 0), fmt_num)
        ws5.write(row5, 6, scheme.get('nlc', 0), fmt_num)
        row5 += 1

print("  ✓ NLC SCHEME REGISTER")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 6: RECONCILIATION SUMMARY (Month-wise CN Comparison)║
# ╚══════════════════════════════════════════════════════════════╝
ws6 = wb.add_worksheet('RECON SUMMARY')
ws6.set_tab_color('#FFC000')
ws6.set_column('A:A', 14)
ws6.set_column('B:B', 16)
ws6.set_column('C:C', 16)
ws6.set_column('D:D', 16)
ws6.set_column('E:E', 16)
ws6.set_column('F:F', 16)
ws6.set_column('G:G', 16)
ws6.set_column('H:H', 14)
ws6.set_column('I:I', 14)
ws6.set_column('J:J', 14)
ws6.set_row(0, 40)

ws6.merge_range('A1:J1', 'RECONCILIATION SUMMARY — EXPECTED vs ACTUAL CREDIT NOTES FY 2025-26', fmt_title)

hdrs6 = ['Month', 'Tally Purchases\n(₹)', 'Stmt Purchases\n(₹)',
          'Expected Secondary\nCN (₹)', 'Expected Primary\nCN (₹)', 'Expected Total\nCN (₹)',
          'Actual CN\nReceived (₹)', 'Difference\n(Act - Exp)(₹)', 'CN Count', 'Status']
ws6.set_row(1, 35)
for ci, h in enumerate(hdrs6):
    ws6.write(1, ci, h, fmt_hdr)

for ri, s in enumerate(monthly_summary):
    r = ri + 2
    ws6.write(r, 0, s['month'], fmt_month)
    ws6.write(r, 1, s['tally_billing_value'], fmt_num)
    ws6.write(r, 2, s['stmt_purchase_value'], fmt_num)
    ws6.write(r, 3, s['exp_secondary_cn'], fmt_num)
    ws6.write(r, 4, s['exp_primary_cn'], fmt_num)
    ws6.write(r, 5, s['exp_total_cn'], fmt_num)
    ws6.write(r, 6, s['actual_cn_received'], fmt_num)
    ws6.write(r, 7, s['cn_difference'], fmt_num)
    ws6.write(r, 8, s['cn_count'], fmt_int)
    sfmt = fmt_green if s['status'] == 'MATCHED' else (fmt_red if s['status'] == 'SHORT' else fmt_orange)
    ws6.write(r, 9, s['status'], sfmt)

gr6 = len(monthly_summary) + 2
ws6.write(gr6, 0, 'FY TOTAL', fmt_grandtot_t)
ws6.write(gr6, 1, sum(s['tally_billing_value'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 2, sum(s['stmt_purchase_value'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 3, sum(s['exp_secondary_cn'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 4, sum(s['exp_primary_cn'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 5, sum(s['exp_total_cn'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 6, sum(s['actual_cn_received'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 7, sum(s['actual_cn_received'] for s in monthly_summary) - sum(s['exp_total_cn'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 8, sum(s['cn_count'] for s in monthly_summary), fmt_grandtot)
ws6.write(gr6, 9, '', fmt_grandtot_t)

# Notes on reconciliation
nr6 = gr6 + 2
ws6.merge_range(nr6, 0, nr6, 9, 'RECONCILIATION NOTES & CAVEATS', fmt_hdr)
recon_notes = [
    ('⚠ IMPORTANT', 'Expected CN is computed from NLC scheme files × Tally quantities. '
     'If your purchase qty in the scheme letter differs from Tally qty, the difference will show here.'),
    ('Scheme Coverage', 'Only products matched to NLC scheme (HIGH/MEDIUM/LOW) are included in expected CN. '
     'UNMATCHED products (not in scheme) have zero expected CN.'),
    ('Aug-2025 Low Billing', 'Tally shows only 50 qty / ₹83K for Aug-2025 vs 284 qty in IML data — '
     'possible missing invoices in Tally or different billing account.'),
    ('Dec-2025', 'Only 5 units in Tally for Dec-2025. Verify if all December billing is captured.'),
    ('ZC Type CNs', 'ZC are non-GST price support credits — treated as CN receipts in statement.'),
    ('Match Confidence', 'LOW confidence matches should be manually verified. Check PRODUCT SCHEME DETAIL sheet.'),
]
for i, (label, note) in enumerate(recon_notes):
    fmt_label = wb.add_format({'bold': True, 'bg_color': '#D6E4BC', 'border': 1, 'text_wrap': True})
    fmt_note  = wb.add_format({'border': 1, 'text_wrap': True})
    ws6.write(nr6 + 1 + i, 0, label, fmt_label)
    ws6.merge_range(nr6 + 1 + i, 1, nr6 + 1 + i, 9, note, fmt_note)
    ws6.set_row(nr6 + 1 + i, 30)

print("  ✓ RECON SUMMARY")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 7: UNMATCHED PRODUCTS                               ║
# ╚══════════════════════════════════════════════════════════════╝
ws7 = wb.add_worksheet('UNMATCHED PRODUCTS')
ws7.set_tab_color('#FF0000')
ws7.set_column('A:A', 14)
ws7.set_column('B:B', 50)
ws7.set_column('C:C', 12)
ws7.set_column('D:D', 50)
ws7.set_row(0, 35)
ws7.merge_range('A1:D1', 'UNMATCHED TALLY PRODUCTS — NO NLC SCHEME FOUND (MANUAL ACTION REQUIRED)', fmt_title)
ws7.write(1, 0, 'Month', fmt_hdr)
ws7.write(1, 1, 'Tally Product Name', fmt_hdr)
ws7.write(1, 2, 'Total Qty', fmt_hdr)
ws7.write(1, 3, 'Action Required', fmt_hdr)

row7 = 2
for month in MONTHS_ORDER:
    results = month_product_scheme.get(month, [])
    for r in results:
        if r['match_conf'] == 'UNMATCHED':
            ws7.write(row7, 0, month, fmt_month)
            ws7.write(row7, 1, r['tally_product'], fmt_text)
            ws7.write(row7, 2, r['qty'], fmt_int)
            ws7.write(row7, 3, 'Verify in NLC scheme letter — No scheme or product name mismatch', fmt_red)
            row7 += 1

print("  ✓ UNMATCHED PRODUCTS")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 8: APPS SCRIPT CODE                                 ║
# ╚══════════════════════════════════════════════════════════════╝
ws8 = wb.add_worksheet('APPS SCRIPT CODE')
ws8.set_tab_color('#4BACC6')
ws8.set_column('A:A', 130)
ws8.merge_range('A1:A2', 'GOOGLE APPS SCRIPT — Fetch IML Credit Notes from Gmail', fmt_title)
ws8.set_row(0, 30)

apps_script = '''/**
 * NALANDA ENTERPRISES — IML Credit Note Gmail Fetcher
 * Paste this in: script.google.com → New Project → Replace code → Save → Run fetchIMLCreditNotes
 * Authorize with: lavesh.bansal@nalandaenterprises.com
 */
function fetchIMLCreditNotes() {
  var ss = SpreadsheetApp.getActiveSpreadsheet();
  var sheet = ss.getSheetByName('CN_From_Gmail') || ss.insertSheet('CN_From_Gmail');
  sheet.clearContents();

  var headers = ['Date', 'Subject', 'From', 'Thread ID', 'Has Attachment',
                 'Attachment Names', 'CN Numbers (from body)', 'Month'];
  sheet.getRange(1, 1, 1, headers.length).setValues([headers]);
  sheet.getRange(1, 1, 1, headers.length).setFontWeight('bold').setBackground('#1F3864').setFontColor('#FFFFFF');

  // Search queries — IML CN emails
  var queries = [
    'from:jitendra.kumar@imaginemarketingindia.com subject:credit note',
    'from:(@imaginemarketingindia.com) subject:"credit note" subject:"nalanda"',
    'from:(@imaginemarketingindia.com) subject:"CN" subject:"nalanda"',
    'from:(@imaginemarketingindia.com) (subject:"scheme" OR subject:"NLC" OR subject:"settlement") subject:"nalanda"',
  ];

  var allThreadIds = new Set();
  var rows = [];
  var cnPattern = /([A-Z]{2,}[\\/\\-]\\d{4,}[\\/\\-]\\w+[\\/\\-]\\d{3,}|CN\\d{6,}|DG\\d{6,})/gi;

  queries.forEach(function(query) {
    var threads = GmailApp.search(query, 0, 200);
    threads.forEach(function(thread) {
      if (allThreadIds.has(thread.getId())) return;
      allThreadIds.add(thread.getId());

      var msgs = thread.getMessages();
      var lastMsg = msgs[msgs.length - 1];
      var date = lastMsg.getDate();
      var month = Utilities.formatDate(date, 'Asia/Kolkata', 'MMM-yyyy');

      var attachments = lastMsg.getAttachments();
      var attachNames = attachments.map(function(a) { return a.getName(); }).join('; ');

      var body = lastMsg.getPlainBody().substring(0, 2000);
      var cnMatches = body.match(cnPattern) || [];
      var cnNumbers = [...new Set(cnMatches)].join(', ');

      rows.push([
        date,
        thread.getFirstMessageSubject(),
        lastMsg.getFrom(),
        thread.getId(),
        attachments.length > 0 ? 'YES' : 'NO',
        attachNames,
        cnNumbers,
        month
      ]);
    });
  });

  if (rows.length > 0) {
    sheet.getRange(2, 1, rows.length, headers.length).setValues(rows);
  }

  SpreadsheetApp.getUi().alert('Done! Found ' + rows.length + ' IML credit note/scheme emails.');
}

/**
 * Fetch only scheme letters (NLC files)
 */
function fetchNLCSchemeEmails() {
  var ss = SpreadsheetApp.getActiveSpreadsheet();
  var sheet = ss.getSheetByName('NLC_Schemes') || ss.insertSheet('NLC_Schemes');
  sheet.clearContents();

  var queries = [
    'from:(@imaginemarketingindia.com) (subject:"NLC" OR subject:"price with scheme" OR subject:"scheme letter")',
    'from:(@imaginemarketingindia.com) (subject:"price list" OR subject:"GT DB" OR subject:"GTDB")',
  ];

  var headers = ['Date', 'Subject', 'From', 'Attachments', 'Month'];
  sheet.getRange(1, 1, 1, headers.length).setValues([headers]);

  var rows = [];
  var seen = new Set();
  queries.forEach(function(q) {
    GmailApp.search(q, 0, 50).forEach(function(t) {
      if (seen.has(t.getId())) return;
      seen.add(t.getId());
      var m = t.getMessages()[0];
      var atts = m.getAttachments().map(a => a.getName()).join('; ');
      var date = m.getDate();
      rows.push([date, t.getFirstMessageSubject(), m.getFrom(), atts,
                 Utilities.formatDate(date, 'Asia/Kolkata', 'MMM-yyyy')]);
    });
  });

  if (rows.length > 0) sheet.getRange(2, 1, rows.length, headers.length).setValues(rows);
  SpreadsheetApp.getUi().alert('Done! Found ' + rows.length + ' NLC scheme emails.');
}

function onOpen() {
  SpreadsheetApp.getUi().createMenu('IML Tools')
    .addItem('Fetch IML Credit Notes', 'fetchIMLCreditNotes')
    .addItem('Fetch NLC Scheme Emails', 'fetchNLCSchemeEmails')
    .addToUi();
}
'''

lines = apps_script.split('\n')
fmt_code = wb.add_format({'font_name': 'Courier New', 'font_size': 9, 'text_wrap': False})
fmt_comment = wb.add_format({'font_name': 'Courier New', 'font_size': 9, 'font_color': '#008000'})
for i, line in enumerate(lines):
    fmt_use = fmt_comment if line.strip().startswith('//') or line.strip().startswith('*') else fmt_code
    ws8.write(i + 3, 0, line, fmt_use)

print("  ✓ APPS SCRIPT CODE")

# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 9: EXECUTIVE SUMMARY                                ║
# ╚══════════════════════════════════════════════════════════════╝
ws9 = wb.add_worksheet('EXECUTIVE SUMMARY')
ws9.set_tab_color('#FFC000')
ws9.set_column('A:A', 35)
ws9.set_column('B:B', 20)
ws9.set_column('C:C', 35)
ws9.set_row(0, 45)
ws9.merge_range('A1:C1', 'EXECUTIVE SUMMARY — NALANDA ENTERPRISES vs IMAGINE MARKETING INDIA FY 2025-26', fmt_title)
ws9.merge_range('A2:C2', f'Prepared: {datetime.now().strftime("%d-%B-%Y")} | Customer Code: 1100164',
                wb.add_format({'italic': True, 'align': 'center', 'bg_color': '#BDD7EE'}))

exec_data = [
    ('PURCHASE OVERVIEW', None, None),
    ('Total Tally Billing (Apr-25 to Mar-26)', f"₹{sum(s['tally_billing_value'] for s in monthly_summary)/1e7:.4f} Cr", 'From Tally Ledger IMAGINE MARKETING LTD MUMBAI 27'),
    ('Total IML Statement Purchases', f"₹{sum(s['stmt_purchase_value'] for s in monthly_summary)/1e7:.4f} Cr", 'From IML SAP Statement (RV type)'),
    ('Total Invoices (Tally)', str(len(invoices)), 'As per Tally billing'),
    ('', None, None),
    ('SCHEME CREDIT NOTES', None, None),
    ('Expected Secondary CN', f"₹{sum(s['exp_secondary_cn'] for s in monthly_summary)/1e5:.2f} L", 'Calculated: NLC secondary rate × Tally qty'),
    ('Expected Primary CN', f"₹{sum(s['exp_primary_cn'] for s in monthly_summary)/1e5:.2f} L", 'Calculated: NLC primary rate × Tally qty'),
    ('Total Expected CN', f"₹{sum(s['exp_total_cn'] for s in monthly_summary)/1e5:.2f} L", 'Secondary + Primary'),
    ('Total Actual CN Received', f"₹{abs(sum(s['actual_cn_received'] for s in monthly_summary))/1e5:.2f} L", 'From IML Statement (DG + CN + ZC)'),
    ('Net Difference (Actual - Expected)', f"₹{(sum(s['actual_cn_received'] for s in monthly_summary) - sum(s['exp_total_cn'] for s in monthly_summary))/1e5:.2f} L", 'Positive = excess received, Negative = short'),
    ('', None, None),
    ('MATCH QUALITY', None, None),
    ('Total Tally Products', str(sum(len(month_product_scheme.get(m,[])) for m in MONTHS_ORDER)), 'Unique product lines across all months'),
    ('Matched (HIGH confidence)', str(sum(1 for m in MONTHS_ORDER for r in month_product_scheme.get(m,[]) if r['match_conf']=='HIGH')), 'Score ≥ 65%'),
    ('Matched (MEDIUM confidence)', str(sum(1 for m in MONTHS_ORDER for r in month_product_scheme.get(m,[]) if r['match_conf']=='MEDIUM')), 'Score 45-65%'),
    ('Matched (LOW confidence)', str(sum(1 for m in MONTHS_ORDER for r in month_product_scheme.get(m,[]) if r['match_conf']=='LOW')), 'Score 35-45% — verify manually'),
    ('UNMATCHED (no scheme found)', str(sum(1 for m in MONTHS_ORDER for r in month_product_scheme.get(m,[]) if r['match_conf']=='UNMATCHED')), 'Not in scheme letter'),
    ('', None, None),
    ('ACTION ITEMS', None, None),
    ('1. Verify LOW confidence matches', 'MANUAL', 'Check PRODUCT SCHEME DETAIL tab, review product names'),
    ('2. Investigate unmatched products', 'MANUAL', 'May be products with no scheme or name variations'),
    ('3. Aug/Dec 2025 billing gap', 'VERIFY', 'Very low Tally billing vs expected — check if all invoices booked'),
    ('4. Cross-check with scheme emails', 'MANUAL', 'Run Apps Script to pull CN emails from Gmail'),
    ('5. Obtain Mar-2026 CN from IML', 'PENDING', 'CN settlement expected in Apr/May 2026'),
]

row9 = 3
fmt_section = wb.add_format({'bold': True, 'bg_color': '#1F3864', 'font_color': '#FFFFFF',
                               'border': 1, 'font_size': 11})
fmt_label9  = wb.add_format({'border': 1, 'bg_color': '#EBF3FB'})
fmt_value9  = wb.add_format({'border': 1, 'bold': True, 'align': 'center', 'bg_color': '#FFF2CC'})
fmt_note9   = wb.add_format({'border': 1, 'italic': True, 'font_color': '#595959'})

for item in exec_data:
    label, value, note = item
    if value is None:
        ws9.merge_range(row9, 0, row9, 2, label, fmt_section if label else fmt_text)
    else:
        ws9.write(row9, 0, label, fmt_label9)
        ws9.write(row9, 1, value, fmt_value9)
        ws9.write(row9, 2, note or '', fmt_note9)
    row9 += 1

ws9.freeze_panes(3, 0)
print("  ✓ EXECUTIVE SUMMARY")

wb.close()
print(f"\n{'='*70}")
print(f"✅ EXCEL FILE CREATED: {OUTPUT_FILE}")
print(f"{'='*70}")

# Final stats
print(f"\nFINAL STATISTICS:")
print(f"  Total Tally Billing:  ₹{sum(s['tally_billing_value'] for s in monthly_summary)/1e7:.4f} Cr")
print(f"  Expected Scheme CN:   ₹{sum(s['exp_total_cn'] for s in monthly_summary)/1e5:.2f} L")
print(f"  Actual CN Received:   ₹{abs(sum(s['actual_cn_received'] for s in monthly_summary))/1e5:.2f} L")
print(f"  Net Difference:       ₹{(sum(s['actual_cn_received'] for s in monthly_summary) - sum(s['exp_total_cn'] for s in monthly_summary))/1e5:.2f} L")
